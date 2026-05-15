import openpyxl
import sqlite3
import datetime
import os
import smtplib
from email.message import EmailMessage
import database

# Email Config (Gmail)
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template.xlsx")

def export_orders_to_excel_and_email(mall_name="현대백화점 대구점"):
    conn = database.get_db()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM settings WHERE id = 1")
    settings = dict(cursor.fetchone())
    
    # Get pending orders and items
    cursor.execute('''
        SELECT o.id, o.order_no, o.customer_name, o.contact, o.address, o.memo, o.created_at,
               i.quantity, p.code, p.name
        FROM orders o
        JOIN order_items i ON o.id = i.order_id
        JOIN products p ON i.product_id = p.id
        WHERE o.status = 'pending'
    ''')
    rows = cursor.fetchall()
    
    if not rows:
        conn.close()
        return {"status": "error", "message": "No pending orders to export."}
    
    # Load Template
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    sheet = wb.active
    
    # Clear existing data starting from row 2 (assuming row 1 is header)
    # Be careful not to delete formatting if possible, but deleting rows is safer
    sheet.delete_rows(2, sheet.max_row)
    
    today_str = datetime.datetime.now().strftime("%Y-%m-%d")
    
    # Map data to excel columns
    for idx, row in enumerate(rows, start=2):
        # 0: 주문일, 1: 몰명, 2: 공급처, 3: 주문번호, 4: 상품코드, 5: 상품명
        # 6: 주문수량, 7: 주문자명, 8: 수취인, 9: 연락처, 10: 주소, 11: 메모, 12: 매출종류
        sheet.cell(row=idx, column=1, value=today_str)
        sheet.cell(row=idx, column=2, value=mall_name) # Dynamic Mall Name
        sheet.cell(row=idx, column=3, value="영업1팀")
        sheet.cell(row=idx, column=4, value=row['order_no'])
        sheet.cell(row=idx, column=5, value=row['code'])
        sheet.cell(row=idx, column=6, value=row['name'])
        sheet.cell(row=idx, column=7, value=row['quantity'])
        sheet.cell(row=idx, column=8, value=row['customer_name'])
        sheet.cell(row=idx, column=9, value=row['customer_name']) # 수취인
        sheet.cell(row=idx, column=10, value=row['contact'])
        sheet.cell(row=idx, column=11, value=row['address'])
        sheet.cell(row=idx, column=12, value=row['memo'])
        sheet.cell(row=idx, column=13, value="신용")
        
    # Save to a new file
    output_filename = f"배송요청_{today_str}.xlsx"
    output_path = os.path.join(os.path.dirname(TEMPLATE_PATH), output_filename)
    wb.save(output_path)
    
    # Mark orders as exported
    order_ids = list(set([r['id'] for r in rows]))
    placeholders = ','.join(['?'] * len(order_ids))
    cursor.execute(f"UPDATE orders SET status = 'exported' WHERE id IN ({placeholders})", order_ids)
    conn.commit()
    conn.close()
    
    # Send Email
    try:
        sender_email = settings.get("sender_email")
        sender_password = settings.get("sender_password")
        receiver_email = settings.get("receiver_email")
        cc_email = settings.get("cc_email")
        sender_name = settings.get("sender_name") or "영업팀"
        
        if not sender_email or not sender_password or not receiver_email:
            return {"status": "error", "message": "이메일 설정이 누락되었습니다. 톱니바퀴 버튼을 눌러 발신자/수신자 설정을 해주세요."}

        msg = EmailMessage()
        msg['Subject'] = f"[백화점 택배주문] {today_str} 배송요청의 건"
        msg['From'] = f"{sender_name} <{sender_email}>"
        msg['To'] = receiver_email
        if cc_email:
            msg['Cc'] = cc_email
        msg.set_content(f"안녕하세요.\n\n{today_str} 백화점 택배 주문건 전달드립니다.\n첨부파일 확인 부탁드립니다.\n\n감사합니다.")
        
        with open(output_path, 'rb') as f:
            excel_data = f.read()
        
        msg.add_attachment(excel_data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=output_filename)
        
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
            
        return {"status": "success", "message": "Exported and emailed successfully."}
    except Exception as e:
        return {"status": "partial_success", "message": f"Excel saved but email failed: {str(e)}"}
